import json
import openpyxl

class Tambola:
    def __init__(self, ticket, name):
        self.ticket = ticket
        self.name = name
        self.all_crossed = False

        self.corners = set()
        self.first_line = set()
        self.second_line = set()
        self.third_line = set()
        self.first_half = set()
        self.second_half = set()
        self.full_house = set()
        self.early_five = set()

        self.find_corner_numbers()
        self.first_line = self.find_line(1)
        self.second_line = self.find_line(2)
        self.third_line = self.find_line(3)
        self.find_first_half()
        self.find_second_half()
        self.find_full_house()
        self.find_early_five()

    def find_line(self, line_num):
        return set([x for x in self.ticket[line_num - 1] if x != 0])
    
    def find_corner_numbers(self):
        for num in self.ticket[0]:
            if num != 0:
                self.corners.add(num)
                break
        for num in self.ticket[0][::-1]:
            if num != 0:
                self.corners.add(num)
                break
        for num in self.ticket[2]:
            if num != 0:
                self.corners.add(num)
                break
        for num in self.ticket[2][::-1]:
            if num != 0:
                self.corners.add(num)
                break

    def find_first_half(self):
        for row in range(3):
            for num in self.ticket[row]:
                if num <= 45 and num != 0:
                    self.first_half.add(num)

    def find_second_half(self):
        for row in range(3):
            for num in self.ticket[row]:
                if num > 45 and num != 0:
                    self.second_half.add(num)

    def find_full_house(self):
        for row in range(3):
            for num in self.ticket[row]:
                if num != 0:
                    self.full_house.add(num)

    def find_early_five(self):
        for row in range(3):
            for num in self.ticket[row]:
                if num != 0:
                    self.early_five.add(num)

    def check_and_cross(self, new_num, early_five_claimed):
        if new_num in self.corners and len(self.corners) > 0:
            self.corners.discard(new_num)
            if len(self.corners) == 0:
                print("Corner - " + str(new_num) + " for " + self.name)

        if new_num in self.first_line and len(self.first_line) > 0:
            self.first_line.discard(new_num)
            if len(self.first_line) == 0:
                print("First Line - " + str(new_num) + " for " + self.name)

        if new_num in self.second_line and len(self.second_line) > 0:
            self.second_line.discard(new_num)
            if len(self.second_line) == 0:
                print("Second Line - " + str(new_num) + " for " + self.name)

        if new_num in self.third_line and len(self.third_line) > 0:
            self.third_line.discard(new_num)
            if len(self.third_line) == 0:
                print("Third Line - " + str(new_num) + " for " + self.name)

        if new_num in self.first_half and len(self.first_half) > 0:
            self.first_half.discard(new_num)
            if len(self.first_half) == 0:
                print("First Half - " + str(new_num) + " for " + self.name)

        if new_num in self.second_half and len(self.second_half) > 0:
            self.second_half.discard(new_num)
            if len(self.second_half) == 0:
                print("Second Half - " + str(new_num) + " for " + self.name)
                
        if new_num in self.full_house and len(self.full_house) > 0:
            self.full_house.discard(new_num)
            if len(self.full_house) == 0:
                self.all_crossed = True
                print("Full house - " + str(new_num) + " for " + self.name)

        if early_five_claimed == 0:
            if new_num in self.early_five and len(self.early_five) > 0:
                self.early_five.discard(new_num)
                if len(self.early_five) == 10:
                    print("Early Five - " + str(new_num) + " for " + self.name)
                    return 1
        
        if early_five_claimed:
            return 1
        else:
            return 0

    def print_all(self):
        print(self.corners)
        print(self.first_line)
        print(self.second_line)
        print(self.third_line)
        print(self.first_half)
        print(self.second_half)
        print(self.full_house)
        print(self.early_five)

def load_data(file_path):
    tickets = {}
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Get All Sheets
    a_sheet_names = wb.sheetnames
    
    # Get Sheet Object by names
    o_sheet = wb[a_sheet_names[0]]

    loop_counter = 0
    for crange in o_sheet.merged_cells:
        clo,rlo,chi,rhi = crange.bounds
        if (chi - clo) == 8 and crange.start_cell.value:
            name = str(crange.start_cell.value)
            rows, cols = (3, 9) 
            numbers = [[0 for i in range(cols)] for j in range(rows)]
            for j in range(3):
                for i in range(9):
                    call_value = o_sheet.cell(row=rlo + j + 1, column=clo + i).value
                    if call_value != "":
                        numbers[j][i] = call_value
                    else:
                        numbers[j][i] = 0
            tickets.update({name:numbers})
 
        loop_counter += 1

    return tickets


import sys
if __name__ == "__main__":
            
    #tambola_obj.print_all()
    
    tickets = load_data("ticket.xlsx")
    tambola_obj_list = []
    
    for ticket in tickets:
        tambola_obj_list.append(Tambola(tickets[ticket], ticket))
        print(ticket, tickets[ticket])

    early_five_claimed = 0
    print("Enter Tambola Number")
    while True:
        input_number = int(input())
        if input_number == -1:
            break
        
        for tambola_obj in tambola_obj_list:
            early_five_claimed = tambola_obj.check_and_cross(input_number, early_five_claimed)
            if tambola_obj.all_crossed == True:
                break
